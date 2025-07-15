"""
    This program is a simple task tracker that allows users to manage tasks using an Excel file.
    It provides functionalities to add, view, filter, update, and export tasks.
    It uses the openpyxl library to handle Excel files.
    The program is designed to be run in a command-line interface (CLI).
    It initializes an Excel file if it does not exist, and allows users to interact with the task tracker through a menu.

"""
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

#Step 0: Initialize the Excel file for task tracking
def initialize_excel(filename='task_tracker.xlsx'):
    """
    Initializes an Excel file for task tracking after checking if it exists.
    If the file does not exist, it creates a new one with the necessary headers.
    """
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks"
        
        #set up headers for the task tracker
        ws.append(["Task ID", "Task Name", "Project", "Date Created", "Due Date", "Status", "Notes"])
        wb.save(filename)
        print(f"Created new task tracker file: {filename}")
    else:
        print(f"Task tracker file already exists: {filename}")

#Step 1: Add a new task to the tracker
def add_task(filename='task_tracker.xlsx'):
    """
    Adds a new task to the task tracker Excel file.
    Prompts the user for task details and appends them to the file.
    """
    wb = load_workbook(filename)
    ws = wb['Tasks']
    
    task_id = ws.max_row  # Use the current row count as the task ID
    task_name = input("Enter task name: ")
    project = input("Enter project name: ")
    date_created = datetime.today().strftime("%Y-%m-%d")
    due_date = input("Enter due date (YYYY-MM-DD): ")
    status = "Pending"  #Defualt status for new tasks
    notes = input("Enter any additional notes: ")
    
    #Append the new task to the worksheet
    ws.append([task_id, task_name, project, date_created, due_date, status, notes])
    wb.save(filename)
    print(f"Task '{task_name}' added successfully.")

#Step 2: View all tasks in the tracker
def view_tasks(filename='task_tracker.xlsx'):
    """
    Displays all tasks from the task tracker Excel file.
    """
    wb = load_workbook(filename)
    ws = wb['Tasks']
    
    print("\n====== All Tasks ======")
    print("-"*60)

    #Iterate from the second row to skip headers
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(row)

#Step 3: Filter tasks based on status
def filter_tasks(filename='task_tracker.xlsx'):
    """
    Filters tasks based on the project name or status keyword.
    """
    filter_type= input("Filter by (project/status): ").strip().lower()
    keyword = input(f"Enter {filter_type} keyword: ").strip()

    wb = load_workbook(filename)
    ws = wb['Tasks']

    print(f"\n Filter Tasks ({filter_type}:{keyword}):")
    print("-"*60)

    for row in ws.iter_rows(min_row=2, values_only=True):
        project=str(row[2]).lower()
        status=str(row[5]).lower()

        # Check if the keyword matches the project or status
        if filter_type == 'project' and keyword in project:
            print(row)
        elif filter_type == 'status' and keyword in status:
            print(row)

#Step 4: Update task status (Placeholder for future implementation)
def update_task_status(filename='task_tracker.xlsx'):
    task_id=int(input("Enter Task ID to update status: "))
    new_status=input("Enter new status (Pending/In Progress/Completed): ").strip()

    wb= load_workbook(filename)
    ws= wb['Tasks']

    found=False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == task_id:  # Check if the Task ID matches
            row[5].value = new_status  # Update the status
            found = True
            print(f"Task ID {task_id} status updated to '{new_status}'.")
            break

    if not found:
        print(f"Task ID {task_id} not found.")

    wb.save(filename)

#Step 5: Export filtered tasks (Placeholder for future implementation)
def export_filtered_tasks(filename='task_tracker.xlsx'):
    """
    Exports filtered tasks to a new Excel file.
    Currently, this function is a placeholder for future implementation.
    """
    export_filename = input("Enter the name for the export file without extension): ").strip()
    keyword=input("Status to export: ").strip().lower()

    wb = load_workbook(filename)
    ws = wb['Tasks']

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Filtered Tasks"

    #Copy headers to the new workbook
    new_ws.append(["Task ID", "Task Name", "Project", "Date Created", "Due Date", "Status", "Notes"])

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[5]).lower() == keyword:  # Check if the status matches
            new_ws.append(row)  # Append the row to the new worksheet
    
    export_file=f"{export_filename}.xlsx"
    new_wb.save(export_file)
    print(f"Filtered tasks exported to {export_file} successfully.")

#Main loop to interact with the user
def menu():
    """
    CLI menu for the task tracker application.
    """
    initialize_excel()
    
    while True:
        print("\n====== Task Tracker Menu ======")
        print("1. Add Task")
        print("2. View all tasks")
        print("3. Filter tasks")
        print("4. Update task status")
        print("5. Export filtered tasks") 
        print("6. Exit")
        
        choice = input("Enter your choice: ").strip() # Get user input for menu choice
        
        if choice == '1':
            add_task()
        elif choice == '2':
            view_tasks()
        elif choice == '3':
            filter_tasks()
        elif choice == '4':
           update_task_status()
        elif choice == '5':
            export_filtered_tasks()
        elif choice == '6':
            print(f"👋 Exiting ... Goodbye!")
            break
        else:
            print("Invalid choice, please try again.")


# Start the script by calling the menu function

if __name__ == "__main__":
    menu()